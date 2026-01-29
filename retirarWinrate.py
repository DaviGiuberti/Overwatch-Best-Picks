#!/usr/bin/env python3
# extract_winrates_two_files.py
# Lê 'winrate1.html' e 'winrate2.html', extrai "name" + "winrate" e salva em 'winrate.xlsx'.
# Coluna A: Hero, B: Winrate (winrate1), C: Winrate (winrate2), D: Average (somente se ambos existirem), E: Saida = 0,2 * Average - 10
# Winrates escritas com vírgula decimal (ex: 47,6). Saida também com vírgula.
# Requer: openpyxl (pip install openpyxl)

import re
import html
import sys
from openpyxl import Workbook

INPUT_HTML1 = "winrate1.html"
INPUT_HTML2 = "winrate2.html"
OUTPUT_XLSX = "winrate.xlsx"

pattern = re.compile(
    r'{"name"\s*:\s*"(?P<name>[^"]+)"[^}]*?"winrate"\s*:\s*(?P<wr>\d+(?:\.\d+)?)',
    re.IGNORECASE | re.UNICODE
)

def parse_file(path):
    """
    Retorna dois dicionários:
     - num_map[name] = float(winrate)  (usado para média)
     - str_map[name] = winrate_com_virgula (ex: '47,6') (usado para escrita no Excel)
    """
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read()
    except FileNotFoundError:
        print(f"Erro: '{path}' não encontrado no diretório atual.")
        sys.exit(1)

    text = html.unescape(raw)
    num_map = {}
    str_map = {}
    for m in pattern.finditer(text):
        name = m.group("name").strip()
        # REMOVER OS ":" DO NOME DO HERÓI
        name = name.replace(":", "")
        name = name.replace(".", "")
        
        wr_text = m.group("wr").strip()  # ex: "47.6"
        try:
            wr_num = float(wr_text)
        except ValueError:
            continue
        # armazenar: número para cálculo e string com vírgula para escrita
        num_map[name] = wr_num
        str_map[name] = wr_text.replace(".", ",")
    return num_map, str_map

def main():
    num1, str1 = parse_file(INPUT_HTML1)    
    num2, str2 = parse_file(INPUT_HTML2)

    # união dos nomes e ordenação alfabética (case-insensitive)
    names_sorted = sorted(set(list(num1.keys()) + list(num2.keys())), key=lambda s: s.lower())

    if not names_sorted:
        print("Nenhuma winrate encontrada nos arquivos. Verifique o conteúdo dos HTMLs.")
        sys.exit(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Winrates"

    # Cabeçalho
    ws["A1"] = "Hero"
    ws["B1"] = "Winrate 1"
    ws["C1"] = "Winrate 2"
    ws["D1"] = "Average"
    ws["E1"] = "Saida (0,2*avg - 10)"

    row = 2
    for name in names_sorted:
        ws.cell(row=row, column=1, value=name)

        # Coluna B: winrate1 (string com vírgula) se existir
        if name in str1:
            ws.cell(row=row, column=2, value=str1[name])
        # Coluna C: winrate2
        if name in str2:
            ws.cell(row=row, column=3, value=str2[name])

        # Coluna D: média — somente se existir em ambos os arquivos
        if name in num1 and name in num2:
            avg = (num1[name] + num2[name]) / 2.0
            # Formatar com 1 casa decimal (ex: 47.6 -> "47,6")
            avg_str = f"{avg:.2f}".replace(".", ",")
            ws.cell(row=row, column=4, value=avg_str)

            # Coluna E: saída = 0.2 * avg - 10
            saida = (0.2 * avg - 10.0) * 2
            saida_str = f"{saida:.2f}".replace(".", ",")
            ws.cell(row=row, column=5, value=saida_str)

        row += 1

    wb.save(OUTPUT_XLSX)
    print(f"OK — {len(names_sorted)} heróis salvos em '{OUTPUT_XLSX}'. Colunas: A=Hero, B=Winrate1, C=Winrate2, D=Average, E=Saida.")

if __name__ == "__main__":
    main()
